import pandas as pd
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook

def generate_staff_excel_template():
    file_path = os.path.join(os.path.dirname(__file__), 'staff_bulk_registration_template.xlsx')
    
    # Define columns
    columns = [
        "primary_psrn_id", "name", "email", "date_of_birth", 
        "gender", "contact_no", "patient_type", "relation", "address"
    ]
    
    # Sample data
    data = [
        ["PSRN-1234", "Dr. John Doe", "john.doe@bits-pilani.ac.in", "1980-05-20", "Male", "9876543210", "Faculty", "", "Quarter No. 42"],
        ["PSRN-1234", "Jane Doe", "jane.doe@gmail.com", "1982-11-15", "Female", "9876543211", "Dependant", "Spouse", "Quarter No. 42"],
        ["PSRN-1234", "Jimmy Doe", "", "2010-02-10", "Male", "", "Dependant", "Son", ""],
        ["PSRN-9999", "Amit Singh", "amit@bits-pilani.ac.in", "1990-01-01", "Male", "8888888888", "Staff", "", "Staff Quarters 10"]
    ]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Registration Template"
    
    # 1. Add Headers
    ws.append(columns)
    
    # Styling Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid") # Blue.800
    alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # 2. Add Sample Data
    for row in data:
        ws.append(row)

    # 3. Data Validation (Drop-downs)
    # Gender Validation
    gender_dv = DataValidation(type="list", formula1='"Male,Female,Other,M,F,O"', allow_blank=True)
    gender_dv.error = 'Select Male (M), Female (F), or Other (O)'
    gender_dv.errorTitle = 'Invalid Gender'
    ws.add_data_validation(gender_dv)
    gender_dv.add('E2:E1000') # Gender is Col E

    # Patient Type Validation
    type_dv = DataValidation(type="list", formula1='"Faculty,Staff,Dependant"', allow_blank=True)
    type_dv.error = 'Your entry is not in the list. Choose Faculty, Staff or Dependant'
    type_dv.errorTitle = 'Invalid Entry'
    ws.add_data_validation(type_dv)
    type_dv.add('G2:G1000') # Patient Type is Col G

    # Relation Validation
    relation_dv = DataValidation(type="list", formula1='"Self,Spouse,Son,Daughter,Father,Mother,Father-in-law,Mother-in-law,Other"', allow_blank=True)
    relation_dv.error = 'Select a valid relation'
    relation_dv.errorTitle = 'Invalid Relation'
    ws.add_data_validation(relation_dv)
    relation_dv.add('H2:H1000') # Relation is Col H

    # 4. Column Widths
    widths = {
        'A': 20, 'B': 25, 'C': 35, 'D': 15, 
        'E': 15, 'F': 15, 'G': 15, 'H': 20, 'I': 30
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # 5. Formatting Date Column (Col D)
    for cell in ws['D'][1:]:
        cell.number_format = 'yyyy-mm-dd'

    # 6. Freeze Top Row
    ws.freeze_panes = "A2"

    # 7. Add Instruction Text on Sheet 2
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["BITS MED-C Bulk Staff & Dependant Registration Guide", ""],
        ["", ""],
        ["1. DO NOT change the header names in the first row.", ""],
        ["2. Sample details have been provided in the first 4 rows. Fill details starting from the fifth row.", ""],
        ["3. Date of Birth must be in YYYY-MM-DD or DD-MM-YYYY format.", ""],
        ["4. Gender, Patient Type, and Relation have drop-down selections for accuracy.", ""],
        ["5. Contact Number should be 10 digits without any spaces or country code.", ""],
        ["6. DEPENDANTS: A Dependant must share the same 'primary_psrn_id' as their Faculty/Staff member.", ""],
        ["7. DEPENDANTS INHERITANCE: If a Dependant leaves 'email', 'contact_no', or 'address' blank, they will automatically inherit those values from the Primary Member.", ""],
        ["8. RELATION: If Patient Type is Faculty or Staff, leave 'relation' blank or select 'Self'. For Dependants, it MUST NOT be 'Self'.", ""],
        ["9. IMPORTANT: When finished, go to File -> Save As -> Choose CSV (.csv) before uploading.", ""],
        ["", ""],
        ["Column Guide:"],
        ["primary_psrn_id *", "PSRN ID of the Faculty/Staff member (e.g., PSRN-1234). Provide this for dependants too. (Mandatory)"],
        ["name *", "Full Name of the patient (Mandatory)"],
        ["email", "Email ID (Mandatory for Primary, Optional for dependants)"],
        ["date_of_birth *", "Format: YYYY-MM-DD or DD-MM-YYYY (Mandatory)"],
        ["gender *", "Select from drop-down or use M, F, O (Mandatory)"],
        ["contact_no", "10-digit mobile number (Mandatory for Primary, Optional for dependants)"],
        ["patient_type *", "Select Faculty, Staff, or Dependant (Mandatory)"],
        ["relation", "Select Self, Son, Daughter, Spouse, etc. (For Faculty/Staff: leave blank or 'Self'. For Dependants: MUST NOT be 'Self')"],
        ["address", "Residential address (Mandatory for Primary, Optional for dependants)"]
    ]
    for row in instructions:
        ws2.append(row)
    
    ws2.column_dimensions['A'].width = 80
    ws2.column_dimensions['B'].width = 80
    ws2['A1'].font = Font(size=14, bold=True)

    wb.save(file_path)
    print(f"Template generated at: {file_path}")

if __name__ == "__main__":
    generate_staff_excel_template()
