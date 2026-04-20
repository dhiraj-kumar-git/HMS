import pandas as pd
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook

def generate_excel_template():
    file_path = os.path.join(os.path.dirname(__file__), 'student_bulk_registration_template.xlsx')
    
    # Define columns
    columns = [
        "institute_id", "name", "email", "date_of_birth", 
        "gender", "contact_no", "patient_type", "address"
    ]
    
    # Sample data
    data = [
        ["2025H1120147P", "Dhiraj Kumar", "h20250147@pilani.bits-pilani.ac.in", "2004-03-15", "Male", "9876543210", "Student", "Ram Bhawan"],
        ["2025H1234567P", "Rahul Sharma", "f20251234@pilani.bits-pilani.ac.in", "2005-08-22", "Male", "9988776655", "Student", "Gandhi Bhawan"],
    ]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Registration Template"
    
    # 1. Add Headers
    ws.append(columns)
    
    # Styling Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid") # Blue.800
    alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # 2. Add Sample Data
    for row in data:
        ws.append(row)

    # 3. Data Validation (Drop-downs)
    # Gender Validation (Accept shorthands)
    gender_dv = DataValidation(type="list", formula1='"Male,Female,Other,M,F,O"', allow_blank=True)
    gender_dv.error = 'Select Male (M), Female (F), or Other (O)'
    gender_dv.errorTitle = 'Invalid Gender'
    ws.add_data_validation(gender_dv)
    gender_dv.add('E2:E1000') # Gender is Col E

    # Patient Type Validation
    type_dv = DataValidation(type="list", formula1='"Student,Faculty,Staff,Dependent,Other"', allow_blank=True)
    type_dv.error = 'Your entry is not in the list'
    type_dv.errorTitle = 'Invalid Entry'
    ws.add_data_validation(type_dv)
    type_dv.add('G2:G1000') # Patient Type is Col G

    # 4. Column Widths
    widths = {
        'A': 20, 'B': 25, 'C': 35, 'D': 15, 
        'E': 15, 'F': 15, 'G': 15, 'H': 30
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # 5. Formatting Date Column (Col D)
    for cell in ws['D'][1:]:
        cell.number_format = 'yyyy-mm-dd'

    # 6. Freeze Top Row
    ws.freeze_panes = "A2"

    # 7. Add Instruction Text on Sheet 2
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["BITS MED-C Bulk Registration Guide"],
        [""],
        ["1. DO NOT change the header names in the first row."],
        ["2. Fill student details starting from the second row."],
        ["3. Date of Birth must be in YYYY-MM-DD format."],
        ["4. Gender and Patient Type have drop-down selections for accuracy."],
        ["5. Contact Number should be 10 digits without any spaces or country code."],
        ["6. IMPORTANT: When finished, go to File -> Save As -> Choose CSV (.csv) before uploading."],
        [""],
        ["Column Guide:"],
        ["institute_id", "Bits ID / Institute ID (e.g., 2025H1120147P)"],
        ["name", "Full Name of the student"],
        ["email", "Official bits-pilani.ac.in email"],
        ["date_of_birth", "Format: YYYY-MM-DD"],
        ["gender", "Select from drop-down or use M, F, O"],
        ["contact_no", "10-digit mobile number"],
        ["patient_type", "Select from drop-down"],
        ["address", "Hostel name or local address"]
    ]
    for row in instructions:
        ws2.append(row)
    
    ws2.column_dimensions['A'].width = 80
    ws2['A1'].font = Font(size=14, bold=True)

    wb.save(file_path)
    print(f"Template generated at: {file_path}")

if __name__ == "__main__":
    generate_excel_template()
